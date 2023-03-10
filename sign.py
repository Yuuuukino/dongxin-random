from openpyxl import *


rooms = ['综408', '综406', '综513', '综508', '综511', '综506', '综413', '综411']
sign_sheet = {}
for i in rooms:
    sign_sheet[i] = []


workbook = load_workbook('新生参赛表.xlsx')
sheet = workbook.active
for sheet in workbook:
    for row in sheet:
        name = row[0].value
        room = row[3].value
        sign_sheet[room].append(name)

workbook = load_workbook('正式参赛表.xlsx')
sheet = workbook.active
for sheet in workbook:
    for row in sheet:
        for j in row:
            print(j.value, end = " ")
        print()
        name = row[0].value
        room = row[3].value
        sign_sheet[room].append(name)


workbook = load_workbook('打星参赛表.xlsx')
sheet = workbook.active
for sheet in workbook:
    for row in sheet:
        for j in row:
            print(j.value, end = " ")
        print()
        name = row[0].value
        room = row[3].value
        sign_sheet[room].append(name)


wb=Workbook()  
ws=wb.active
for i in sign_sheet:    
    sheet = wb.create_sheet(i)
    for j in sign_sheet[i]:
        sheet.append([j])
wb.remove(wb._sheets[0])

wb.save('签到表.xlsx')
