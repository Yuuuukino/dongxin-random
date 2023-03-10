from random import *
from os import*
from openpyxl import *

# 导入
# path = r"C:/Users/Sakana/Desktop"
path = getcwd()
chdir(path)
workbook = load_workbook('报名表.xlsx')
sheet = workbook.active
LIST_New = []
LIST_Old = []
LIST_Star = []

# 将数据放入列表
for i in range(2, sheet.max_row + 1):
    cell_name = sheet.cell(row = i, column = 3)
    cell_xuehao = sheet.cell(row = i, column = 2)
    cell_group = sheet.cell(row = i, column = 11)
    cell_xueyuan = sheet.cell(row = i, column = 5)  
    cell_star = sheet.cell(row = i, column = 10)
    cell_xuexiao = sheet.cell(row = i, column = 6)
    tmp = [cell_name.value, cell_xuehao.value, cell_xueyuan.value]
    if cell_star.value == '打星':
        tmp[2] = cell_xuexiao.value
        LIST_Star.append(tmp)
    elif cell_group.value == '新生组':
        LIST_New.append(tmp)
    else:
        LIST_Old.append(tmp)
        

# 乱序，输出到新excel
shuffle(LIST_New)
shuffle(LIST_Old)
shuffle(LIST_Star)
print(len(LIST_New), len(LIST_Old), len(LIST_Star))


# 新生
# 将编号放入新表
workbook = load_workbook('新生组.xlsx')
Sheet_names = workbook.sheetnames
LIST_TOT_NEW = []
for sheet in workbook:
    Classroom = Sheet_names[0]
    del Sheet_names[0]
    x = (sheet.dimensions).split(":")
    for row in sheet:
        for column in row:
            if len(LIST_New) == 0:
                break
            if column.value == 1:
                pos = column.coordinate
                LIST_New[-1].append(str(Classroom))
                LIST_New[-1].append(str(pos))
                tmp = "&".join([str(x) for x in LIST_New.pop()])
                # sheet[pos] = " ".join(tmp.split("-"))
                LIST_TOT_NEW.append(tmp)
                tmp = tmp.split("&")
                sheet[pos] = tmp[-3]
workbook.save('新生组_座位表.xlsx')
workbook = Workbook()
sheet = workbook.active
for i in LIST_TOT_NEW:
    sheet.append(i.split("&"))
workbook.save('新生参赛表.xlsx')


# 老生
# 将编号放入新表
workbook = load_workbook('正式组.xlsx')
Sheet_names = workbook.sheetnames
LIST_TOT_OLD = []
for sheet in workbook:
    Classroom = Sheet_names[0]
    del Sheet_names[0]
    x = (sheet.dimensions).split(":")
    for row in sheet:
        for column in row:
            if len(LIST_Old) == 0:
                break
            if column.value == 1:
                pos = column.coordinate
                LIST_Old[-1].append(str(Classroom))
                LIST_Old[-1].append(str(pos))
                tmp = "&".join([str(x) for x in LIST_Old.pop()])
                # sheet[pos] = " ".join(tmp.split("-"))
                LIST_TOT_OLD.append(tmp)
                tmp = tmp.split("&")
                sheet[pos] = tmp[-3]
workbook.save('正式组_座位表.xlsx')
workbook = Workbook()
sheet = workbook.active
for i in LIST_TOT_OLD:
    sheet.append(i.split("&"))
workbook.save('正式参赛表.xlsx')


# 打星
# 将编号放入新表
workbook = load_workbook('正式组_座位表.xlsx')
Sheet_names = workbook.sheetnames
LIST_TOT_STAR = []
for sheet in workbook:
    Classroom = Sheet_names[0]
    del Sheet_names[0]
    x = (sheet.dimensions).split(":")
    for row in sheet:
        for column in row:
            if len(LIST_Star) == 0:
                break
            if column.value == 1:
                pos = column.coordinate
                LIST_Star[-1].append(str(Classroom))
                LIST_Star[-1].append(str(pos))
                tmp = "&".join([str(x) for x in LIST_Star.pop()])
                # sheet[pos] = " ".join(tmp.split("-"))
                LIST_TOT_STAR.append(tmp)
                tmp = tmp.split("&")
                sheet[pos] = tmp[-3]
workbook.save('正式组_座位表.xlsx')
workbook = Workbook()
sheet = workbook.active
for i in LIST_TOT_STAR:
    sheet.append(i.split("&"))
workbook.save('打星参赛表.xlsx')

